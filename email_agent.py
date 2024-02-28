# -*- coding: utf-8 -*-
"""
Created on Sun Feb 11 20:41:01 2024

@author: user
"""

# Imports
# 1. GUI
import tkinter as tk
from tkinter.filedialog import askopenfilename
from tkinter import ttk, scrolledtext, Menu

# 2. Server communication
import smtplib # libraria pentru trimitere e-mail
from email.mime.multipart import MIMEMultipart # libraria pentru formatarea mesajului
from email.mime.text import MIMEText # pentru textul scrisorii
from email.mime.application import MIMEApplication # pentru attachment

# 3. Other libs
import openpyxl, re, os, time, winsound, webbrowser, sys, git
from PIL import Image


""" ************************** Main Functions ******************************"""


""" ************************** Back END  ***********************************"""

""" ************************** Server communication ************************"""

def connect_server(from_addr, password):
    """ Functia care stabileste conexiunea cu serverul"""
    global server
    # Vom verifica ce fel de posta utilizeaza user-ul
    tail = from_addr.split("@")[-1] # coada adresei (yahoo.com, gmail.com, mt.utm.md, mail.ru)
    flag = None
    if tail == "yahoo.com":
        server_adress = 'smtp.mail.yahoo.com'; server_port = 465
        flag = 1
        
    elif tail == "gmail.com":
        server_adress = 'smtp.gmail.com'; server_port = 465
        flag = 1
    
    elif tail == "mail.ru":
        server_adress = 'smtp.mail.ru'; server_port = 465
        flag = 1

    else:
        info_msg = "Your email is not available now, please contact the developper"
        popup_msg(info_msg)
        flag = 0
        
    server = None
    if flag == 1: # Pentru alte servere de mail
        try:
            server = smtplib.SMTP_SSL(server_adress, server_port)
        except:
            popup_msg("Failed to connect server. Check your login or internet connection")
        try:
            server.ehlo() # Trimitem un fel de hello serverului, sa vedem daca raspunde
        except:
            popup_msg("Server not responding ")

            
    # Daca am reusit conexiunea cu serverul, atunci mergem mai departe
    if server:
        try:
            server.login(from_addr, password) # Ne logam pe account-ul nostru
            check_server_connection_button.configure(bg='light blue',font=fnt, text = "Connected...", state = 'disabled')
            
            send_mail_to_yourself_button.configure(state = 'normal', bg = 'yellow')
            send_button.configure(state = 'normal', bg = 'green')
            sender_email_entry.configure(state='disabled')
            sender_password_entry.configure(state='disabled')
        except:
            popup_msg("Failed to login. Check your login and password")


""" ************************** Email Extractor **********************************"""

def open_excel(path_to_excel):
    try:
        print("Reading excel file ...")
        t1 = time.time()
        excel_obj = openpyxl.load_workbook(path_to_excel, read_only=True, data_only=True)
        t2 = time.time()
        print("Done in {t:.1f} sec.\n".format(t = t2-t1))
    except:
        popup_msg("Unable to read excel file. \ Please close the program and try again")
    # Then, return the excel_object
    return excel_obj

def read_excel(excel_obj):
    # Extracting sheets from excel object
    sheets = excel_obj.worksheets
    # Extracting sheet names (only to be printed in console)
    sheet_names = excel_obj.sheetnames
    # Prepare an empty list to be filled with content extracted from sheets
    emails = []
    # Initiate a counter, to count sheets extracted (only to print in console)
    sheet_counter = 0
    # Initiate timer, to print total extracting time in console
    # Declare regex pattern for emails
    regex = re.compile(r"([-!#-'*+/-9=?A-Z^-~]+(\.[-!#-'*+/-9=?A-Z^-~]+)*|\"([]!#-[^-~ \t]|(\\[\t -~]))+\")@([-!#-'*+/-9=?A-Z^-~]+(\.[-!#-'*+/-9=?A-Z^-~]+)*|\[[\t -Z^-~]*])") 
    t1 = time.time()
    print("Start extracting cell values from every sheet")
    for sheet in sheets:
        t11 = time.time()
        # Get maximum number of rows with data in every sheet
        rows = 0
        for max_row, row in enumerate(sheet,1):
            if not all(col.value is None for col in row):
                rows +=1
        # Extract all values from indicated columns
        print(f"Extracting cell values from sheet {sheet_names[sheet_counter]} ({rows} rows)")
        for row in sheet.iter_rows(min_row=1, min_col = 1, max_row = rows, max_col = 11):
            for cell in row:
                cell_val = str(cell.value).split("\n")
                for i in cell_val:
                    if re.fullmatch(regex, i.strip()):
                       emails.append(i.strip())
        t22 = time.time()
        print("Done in {t:.1f} sec.".format(t = t22-t11))
        sheet_counter+=1
    #garbage collector
    excel_obj.close()
    # Removing duplicates from list
    # ! Set is a collection of unique items
    emails = set(emails)
    # Generate a sound when finishes
    freq = 1250
    dur = 100
    for i in range(0, 5):    
        winsound.Beep(freq, dur)    
        freq+= 100   
    t2 = time.time()
    print("\n All cells from columns A to K, from all sheets were read in {t:.1f} sec.".format(t = t2-t1))
    print(f"\n Exporting {len(emails)} emails... \n")
    # Call function export emails
    t1 = time.time()
    export_emails(emails)
    t2 = time.time()
    print("\n Done, emails are updated! in {t:.1f} sec.".format(t = t2-t1))
    
    
def export_emails(path_to_excel, emails):
    # check if there exists an excel book for emails:
    # A little modification of path,in order to point to new exmails.xlsx document
    dir_path = os.path.dirname(path_to_excel)
    export_path = os.path.abspath(f"{dir_path}/emails.xlsx")
    if not os.path.isfile(export_path):
        # create a new workbook
        wb = openpyxl.Workbook()
        sheet = wb.active
        for i in range(len(emails)):
            #Cell object is created by  using sheet object's cell() method.
            c = sheet.cell(row = i+1, column = 1) 
            # writing values to cells 
            c.value = list(emails)[i]
        wb.save(export_path)
        # garbage collector
        wb.close()
    else:
        # if the file already exists, read emails and update list
        # Read excel object
        excel_obj = openpyxl.load_workbook(export_path,read_only=True)
        sheet = excel_obj.worksheets[0]
        # Get max num of rows
        rows = 0
        for max_row, row in enumerate(sheet,1):
            if not all(col.value is None for col in row):
                rows +=1
        # Extract all values from indicated columns
        old_emails = []
        for row in sheet.iter_rows(min_row=1, min_col = 1, max_row = rows, max_col = 1):
            for cell in row:
                cell_val = cell.value 
                if cell_val is not None: # Remove none from list, if exists
                    old_emails.append(cell_val)
        # Now, we will join emails list with old_emails and will remove duplicates
        new_emails = set(old_emails + list(emails))
        # And we will rewrite all into the same excel file
        wb = openpyxl.Workbook()
        sheet = wb.active
        for i in range(len(new_emails)):
            #Cell object is created by  using sheet object's cell() method.
            c = sheet.cell(row = i+1, column = 1) 
            # writing values to cells 
            c.value = list(new_emails)[i]
        wb.save(export_path)
        # Garbage collector
        excel_obj.close()
        wb.close()

def read_excel_with_emails(excel_obj):
    # Read excel with emails
    sheet = excel_obj.worksheets[0]
    # Prepare an empty list to be filled with content extracted from sheets
    recipient_emails = []
    # Get max num of rows
    rows = 0
    for max_row, row in enumerate(sheet,1):
        if not all(col.value is None for col in row):
            rows +=1
    # Declare regex pattern for emails
    regex = re.compile(r"([-!#-'*+/-9=?A-Z^-~]+(\.[-!#-'*+/-9=?A-Z^-~]+)*|\"([]!#-[^-~ \t]|(\\[\t -~]))+\")@([-!#-'*+/-9=?A-Z^-~]+(\.[-!#-'*+/-9=?A-Z^-~]+)*|\[[\t -Z^-~]*])")
    for row in sheet.iter_rows(min_row=1, min_col = 1, max_row = rows, max_col = 1):
        for cell in row:
            cell_val=str(cell.value)
            if re.fullmatch(regex, cell_val):
               recipient_emails.append(cell_val)
    return recipient_emails
               
               
""" ************************** Email Sender ********************************"""
def format_message_content(from_addr, to_addr, subject, content):
    """ Functia care formeaza continutul e-mail-ului"""
    # Mesajul inclus in email:
    msg = MIMEMultipart() # cream obiectul msg
    # 1. Antetetul scrisorii
    msg['Subject'] = subject
    msg['From'] = from_addr
    msg['To'] = to_addr
    # 2. Continutul de text al scrisorii
    body = MIMEText(content, 'html') # creaza obiectul html - body
    msg.attach(body)
    # 3. Attachment la scrisoare
    if attachment_path:
        file_to_attach = MIMEApplication(open(attachment_path, "rb").read())
        # 3.3. Indicam concret tipul de attachment:
        # ! Vom extrage numele fisierului pentru a fi inclus in filename
        file_to_attach.add_header('Content-Disposition', 'attachment',\
                                    filename=os.path.basename(attachment_path))
        #    unde: add_header(_name, _value, **_params) - 'Content-Disposition' indica
        #          codului htmp cum va fi afisat contentul. In cazul dat ca attachment
        # 3.4. Atasam si aceasta la mesaj:
        msg.attach(file_to_attach)
    return msg

""" ************************** GUI  ********************************"""

""" GUI-linked functions"""

def browse_excel():
    global excel_obj
    global path_to_excel
    path_to_excel = askopenfilename(title="Select a File",\
                                 filetype=(("Excel", "*.xlsx"), ("Excel", "*.xls"))) 
    # Transmitem path-ul catre open_excel
    excel_obj = open_excel(path_to_excel)
    
    
def preview():
    # Now, we have excel_obj, path_to_excel and sheet_name
    with open(f"{os.path.dirname(path_to_excel)}\\{os.path.basename(path_to_excel).split('.')[:-1][0]}_Preview.html", mode = 'w', encoding="utf-8") as file:
        file.write('Informatia care va fi expediata la fiecare adresant: <br>')
        file.write(f"Subiect: {str(subject_entry.get())}  <br>")
        content = "<br> ".join(message_content_entry.get('1.0','end-1c').split('\n'))
        file.write(f"Continut: <br> {content} <br>")
        if attachment_path:
            with Image.open(attachment_path) as img:
                width, height = (img.size)
                html_image_code = f"<img src='{attachment_path}' alt='image' width='{width}' height='{height}'>"
            file.write(html_image_code)
    url = f"{os.path.dirname(path_to_excel)}\\{os.path.basename(path_to_excel).split('.')[:-1][0]}_Preview.html"
    new = 2 # open in a new tab, if possible
    webbrowser.open(url,new=new)
    
    
    
def send_mail_to_yourself():
    """Functia pentru a expedia mesaj pe propriul account"""
    try:
        msg = format_message_content(sender_email_entry.get(), sender_email_entry.get(), subject_entry.get(), message_content_entry.get('1.0','end-1c'))
    except:
        popup_msg("Verificati datele introduse in campuri")
    try:
        server.sendmail(sender_email_entry.get(), sender_email_entry.get(), msg.as_string()) # Trimitem e-mail de pe account-ul nostru
        popup_msg('Pe adresa Dvs. a fost expediat un email de testare') 
    except:
        popup_msg("Connection with mail server lost")


def send_mail_to_recipients():
    """Functia pentru a expedia mesaje la fiecare email din excel"""
    recipient_emails = read_excel_with_emails(excel_obj)
    msg = format_message_content(sender_email_entry.get(), sender_email_entry.get(), subject_entry.get(), message_content_entry.get('1.0','end-1c'))
    for item in recipient_emails:
        to_addr = item
        try:
            to_addr = item
            server.sendmail(sender_email_entry.get(), to_addr, msg.as_string())
            print(f"Successfully sent to {item}")
        except:
            print(f"Message not sent to {item}")
            pass
    popup_msg("Finished sending emails...")


def set_attachment():
    global attachment_path
    attachment_path = os.path.abspath(askopenfilename(title="Select a File",\
                                 filetype=(('JPG Files', '*.jpg'),
                                           ('JPEG Files', '*.jpeg'),
                                           ('PNG Files', '*.png'),
                                           ('All files','*.*'))))

def popup_msg(message):
    popup = tk.Tk()
    popup.wm_title("Info")
    label = tk.Label(popup, text=message, fg = 'blue' )
    label.pack(side="top", fill="x", pady=20)
    label.config(font=("Times New Roman", 14))
    B1 = tk.Button(popup, text="Ok", padx = 30, pady = 5, borderwidth = 5, \
                   command = popup.destroy)
    B1.pack()


def exit_app():
    # Garbage collector
    # Close all connections
    try:
        if excel_obj:
            excel_obj.close()
    except:
        pass
    for obj in dir():
        del globals()[obj]
    root.destroy()

def check_for_updates():
    parent_path = os.path.abspath(os.path.join(sys.executable, os.pardir))
    repo_path = os.path.abspath(os.path.join(parent_path,"MyScripts\\Email-Agent" ))
    repo = git.Repo(repo_path)
    repo.remotes.origin.fetch()
    diff = repo.git.diff('origin/main')
    if len(diff) !=0:
        print('diff = ', diff)
        popup_msg("New updates are avaialble!")
        
def update():
    parent_path = os.path.abspath(os.path.join(sys.executable, os.pardir))
    repo_path = os.path.abspath(os.path.join(parent_path,"MyScripts\\Email-Agent" ))
    repo = git.Repo(repo_path)
    repo.git.reset('--hard','origin/main')
    origin = repo.remote(name='origin')
    origin.fetch()
    diff = repo.git.diff('origin/main')
    if len(diff) !=0:
        origin.pull()
        popup_msg("Successfully updated. Please restart the program!")
    else:
        popup_msg("Your program is up to date")


""" GUI graphical elements """

root = tk.Tk()
root.title("Email Agent")
root.geometry("850x600")

# When start GUI, we'll check for updates
check_for_updates()

# Some global variables...
attachment_path = None


# Tabs
tab1 = ttk.Frame
tabControl = ttk.Notebook(root) 
  
tab1 = ttk.Frame(tabControl) 
tab2 = ttk.Frame(tabControl) 
  
tabControl.add(tab1, text ='Email Sender') 
tabControl.add(tab2, text ='Email Extractor') 
tabControl.grid(row = 0, column = 0)

# Font specifications
fnt = ("Arial", 12, "bold")

# Menu
menubar = Menu(root) # generam un obiect menubar in root
filemenu = Menu(menubar, tearoff=0) # Definim obiectul pentru File
filemenu.add_command(label="New", command=lambda: popup_msg('under process...')) # Adaugam comanda New
filemenu.add_separator()
filemenu.add_command(label = 'Exit', command = exit_app)
menubar.add_cascade(label="File", menu=filemenu) #In bara de meniuri includem obiectul filemenu

editmenu = Menu(menubar, tearoff=0)
editmenu.add_command(label="Delete", command=lambda: popup_msg('under process...'))
menubar.add_cascade(label="Edit", menu=editmenu)

helpmenu = Menu(menubar, tearoff=0)
helpmenu.add_command(label="Update", command=update)
helpmenu.add_command(label="About...", command=lambda: popup_msg('A simple application for extracting and sending emails'))
menubar.add_cascade(label="Help", menu=helpmenu)

root.config(menu = menubar) # Configuram root-ul ca sa stie ca in menu avem obiectul menubar


# GUI elements definition, Email Sender
browse_excel_label = tk.Label(tab1, text = "Excel file with email adresses   ",font = fnt)
browse_excel_button = tk.Button(tab1, text = "browse", padx = 30, pady = 5,borderwidth = 5, font=fnt,\
               command = browse_excel)

    
sender_label = tk.Label(tab1, text = "Sender email", font=fnt)
sender_email_entry = tk.Entry(tab1,  width = 50, borderwidth = 5, font=fnt)
# Temporar, inseram un email de testare
sender_email_entry.insert(0,'')
    
sender_password_label = tk.Label(tab1, text = "Sender password", font=fnt)
sender_password_entry = tk.Entry(tab1,  width = 50, borderwidth = 5, font=fnt, show = "*")
# Temporar, inseram parola
sender_password_entry.insert(0,'')
    
subject_label = tk.Label(tab1, text = "Subject", font=fnt)
subject_entry = tk.Entry(tab1, width = 50, borderwidth = 5,font = fnt)
subject_entry.insert(tk.END,'Please, consider this message')

message_content_label = tk.Label(tab1, text = "Message content ", font=fnt)
message_content_entry = scrolledtext.ScrolledText(tab1, wrap = tk.WORD, height = 12, width = 50,  borderwidth = 5, font=fnt)

message_content_entry.insert('1.0','Content pentru testare')


attachment_label = tk.Label(tab1, text = "Want to attach some images? ", font=fnt)
attachment_browse_button = tk.Button(tab1, text = "browse", padx = 30, pady = 5,borderwidth = 5, font=fnt,\
               command = set_attachment)
# Buttons
check_server_connection_button = tk.Button(tab1, text = "Check \nserver \nconnection ", fg = 'black', font=fnt,\
              padx = 10, pady = 5, borderwidth = 5,\
              bg = 'light blue',command = lambda: connect_server(sender_email_entry.get(), sender_password_entry.get()))

preview_button = tk.Button(tab1, text = "Preview", fg = 'black', font=fnt,\
              padx = 25, pady = 8, borderwidth = 5,\
              bg = 'yellow',command = preview)
    
send_mail_to_yourself_button = tk.Button(tab1, state = 'disabled', text = "Send \ntest email", fg = 'black', font=fnt,\
              padx = 20, pady = 8, borderwidth = 5,\
              bg = 'gray',command = send_mail_to_yourself)


send_button = tk.Button(tab1, text = "SEND",fg = 'black',font=fnt,\
               padx = 33, pady = 20, borderwidth = 5,\
               bg = 'gray',command = send_mail_to_recipients)

    
exit_button = tk.Button(tab1, text = "EXIT", fg = 'black', font=fnt,\
              padx = 37, pady = 10, borderwidth = 5,\
              bg = 'red',command =  exit_app) 
    
    
# GUI elements definition, Email Extractor

browse_excel_with_emails_title_label = tk.Label(tab2, text = "Select an excel file to scan and extract emails",font = fnt)

browse_excel_with_emails_label = tk.Label(tab2, text = "Excel file: ",font = fnt)
browse_excel_with_emails_button = tk.Button(tab2, text = "browse", padx = 30, pady = 2,borderwidth = 5, font=fnt,\
               command = browse_excel)

"""************* Window layout **************"""

# Email Sender
r = 0
browse_excel_label.grid(sticky = "W", row = r, column = 0, columnspan = 4)
browse_excel_button.grid(sticky = "W", row = r, column = 4, columnspan = 2)

r+=1
sender_label.grid(sticky = "W", row = r, column = 0, columnspan = 2)
sender_email_entry.grid(sticky = "W",row = r, column = 3, columnspan = 2)
check_server_connection_button.grid(row = r, column = 6, columnspan = 2, rowspan = 2)

r+=1
sender_password_label.grid(row = r, column = 0, columnspan = 2)
sender_password_entry.grid(sticky = "W",row = r, column = 3, columnspan = 2)

r+=1
subject_label.grid(sticky = "W",row = r, column = 0, columnspan = 2)
subject_entry.grid(sticky = "W",row = r, column = 3, columnspan = 4)
preview_button.grid(row = r, column = 6, columnspan = 2)

r+=1
message_content_label.grid(sticky = "W",row = r, column = 0, columnspan = 2)
message_content_entry.grid(sticky = "W",row = r, column = 3, columnspan = 6, rowspan = 6)
send_mail_to_yourself_button.grid(sticky = "N", row = r, column = 6, columnspan = 2)
r+=1
send_button.grid(sticky = "N",row = r, column = 6, columnspan = 2)
r+=1
exit_button.grid(sticky = "S",row = r, column = 6, columnspan = 2)

r+=7
attachment_label.grid(sticky = "W",row = r, column = 0, columnspan = 2)
attachment_browse_button.grid(sticky = "W",row = r, column = 3, columnspan = 6)

# Email Extractor
browse_excel_with_emails_title_label.grid(sticky="W", row = 0, column = 0, columnspan = 4)
browse_excel_with_emails_label.grid(sticky="E", row = 1, column = 0, rowspan = 2)
browse_excel_with_emails_button.grid(sticky="E",row = 1, column = 2, rowspan = 2)    




root.mainloop()
    
    
